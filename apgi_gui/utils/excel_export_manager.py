"""
Excel Export Manager for APGI Framework GUI

Provides comprehensive Excel export functionality for experimental data, results, and reports.
Supports multiple data formats and customizable styling.
"""

import os
from datetime import datetime
from typing import Any, Dict, List, Union

import pandas as pd

try:
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, Reference
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils.dataframe import dataframe_to_rows

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print(
        "Warning: openpyxl not available. Excel export functionality will be limited."
    )

try:
    import xlsxwriter  # noqa: F401

    XLSXWRITER_AVAILABLE = True
except ImportError:
    XLSXWRITER_AVAILABLE = False


class ExcelExportManager:
    """Manages Excel export functionality for GUI applications."""

    def __init__(self):
        self.default_styles = self._create_default_styles()

    def _create_default_styles(self) -> Dict[str, Any]:
        """Create default Excel styles."""
        if not OPENPYXL_AVAILABLE:
            return {}

        return {
            "header_font": Font(bold=True, color="FFFFFF"),
            "header_fill": PatternFill(
                start_color="366092", end_color="366092", fill_type="solid"
            ),
            "header_alignment": Alignment(horizontal="center", vertical="center"),
            "data_font": Font(size=11),
            "data_alignment": Alignment(horizontal="left", vertical="center"),
            "border": Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            ),
        }

    def export_to_excel(
        self,
        data: Union[Dict, List, pd.DataFrame],
        filepath: str,
        sheet_name: str = "Data",
        include_charts: bool = False,
        styling: bool = True,
    ) -> bool:
        """Export data to Excel format."""
        try:
            if OPENPYXL_AVAILABLE:
                return self._export_with_openpyxl(
                    data, filepath, sheet_name, include_charts, styling
                )
            elif XLSXWRITER_AVAILABLE:
                return self._export_with_xlsxwriter(
                    data, filepath, sheet_name, include_charts
                )
            else:
                # Fallback to pandas
                return self._export_with_pandas(data, filepath, sheet_name)
        except Exception as e:
            print(f"Error exporting to Excel: {e}")
            return False

    def _export_with_openpyxl(
        self,
        data: Union[Dict, List, pd.DataFrame],
        filepath: str,
        sheet_name: str,
        include_charts: bool,
        styling: bool,
    ) -> bool:
        """Export using openpyxl with advanced formatting."""
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        # Convert data to DataFrame if needed
        if isinstance(data, dict):
            df = pd.DataFrame([data])
        elif isinstance(data, list):
            df = pd.DataFrame(data)
        else:
            df = data

        # Write headers
        for col_num, column in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col_num, value=column)
            if styling:
                cell.font = self.default_styles["header_font"]
                cell.fill = self.default_styles["header_fill"]
                cell.alignment = self.default_styles["header_alignment"]
                cell.border = self.default_styles["border"]

        # Write data
        for row_num, row in enumerate(
            dataframe_to_rows(df, index=False, header=False), 2
        ):
            for col_num, value in enumerate(row, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                if styling:
                    cell.font = self.default_styles["data_font"]
                    cell.alignment = self.default_styles["data_alignment"]
                    cell.border = self.default_styles["border"]

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter  # type: ignore
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except (AttributeError, ValueError):
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Add charts if requested
        if include_charts and len(df) > 1:
            self._add_charts(ws, df)

        # Save workbook
        wb.save(filepath)
        return True

    def _export_with_xlsxwriter(
        self,
        data: Union[Dict, List, pd.DataFrame],
        filepath: str,
        sheet_name: str,
        include_charts: bool,
    ) -> bool:
        """Export using xlsxwriter with advanced features."""
        # Convert data to DataFrame
        if isinstance(data, dict):
            df = pd.DataFrame([data])
        elif isinstance(data, list):
            df = pd.DataFrame(data)
        else:
            df = data

        # Create writer
        with pd.ExcelWriter(filepath, engine="xlsxwriter") as writer:
            # Write data
            df.to_excel(writer, sheet_name=sheet_name, index=False)

            # Get workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets[sheet_name]

            # Add formats
            header_format = workbook.add_format(
                {
                    "bold": True,
                    "text_wrap": True,
                    "valign": "top",
                    "fg_color": "#366092",
                    "font_color": "white",
                    "border": 1,
                }
            )

            data_format = workbook.add_format(
                {"border": 1, "text_wrap": True, "valign": "top"}
            )

            # Apply formatting to headers
            for col_num, value in enumerate(df.columns):
                worksheet.write(0, col_num, value, header_format)

            # Apply formatting to data
            for row_num in range(1, len(df) + 1):
                for col_num in range(len(df.columns)):
                    worksheet.write(
                        row_num, col_num, df.iloc[row_num - 1, col_num], data_format
                    )

            # Auto-adjust column widths
            for col_num, column in enumerate(df.columns):
                max_len = (
                    max(df[column].astype(str).map(len).max(), len(str(column))) + 2
                )
                worksheet.set_column(col_num, col_num, max_len)

            # Add charts if requested
            if include_charts and len(df) > 1:
                self._add_xlsxwriter_charts(worksheet, workbook, df, len(df.columns))

        return True

    def _export_with_pandas(
        self, data: Union[Dict, List, pd.DataFrame], filepath: str, sheet_name: str
    ) -> bool:
        """Fallback export using pandas only."""
        try:
            # Convert data to DataFrame
            if isinstance(data, dict):
                df = pd.DataFrame([data])
            elif isinstance(data, list):
                df = pd.DataFrame(data)
            else:
                df = data

            # Export using pandas
            df.to_excel(filepath, sheet_name=sheet_name, index=False)
            return True
        except Exception as e:
            print(f"Error with pandas export: {e}")
            return False

    def _add_charts(self, worksheet, df: pd.DataFrame) -> None:
        """Add charts to the worksheet using openpyxl."""
        if len(df.columns) < 2:
            return

        # Create a bar chart for numeric columns
        numeric_cols = df.select_dtypes(include=["number"]).columns
        if len(numeric_cols) > 0:
            chart = BarChart()
            chart.type = "col"
            chart.style = 10
            chart.title = "Data Visualization"
            chart.y_axis.title = "Values"
            chart.x_axis.title = "Categories"

            # Add data
            data = Reference(
                worksheet,
                min_col=1,
                min_row=1,
                max_col=len(numeric_cols),
                max_row=len(df) + 1,
            )
            cats = Reference(
                worksheet, min_col=1, min_row=2, max_col=1, max_row=len(df) + 1
            )

            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)

            # Position chart
            worksheet.add_chart(chart, "A20")

    def _add_xlsxwriter_charts(
        self, worksheet, workbook, df: pd.DataFrame, num_cols: int
    ) -> None:
        """Add charts using xlsxwriter."""
        numeric_cols = df.select_dtypes(include=["number"]).columns
        if len(numeric_cols) > 0:
            # Create a chart
            chart = workbook.add_chart({"type": "column"})

            # Get sheet name for chart references
            sheet_name = worksheet.name

            # Configure chart
            chart.add_series(
                {
                    "name": "Data",
                    "categories": [sheet_name, 1, 1, len(df), 1],
                    "values": [sheet_name, 1, 2, len(df), 2],
                }
            )

            chart.set_title("Data Visualization")
            chart.set_x_axis({"name": "Categories"})
            chart.set_y_axis({"name": "Values"})

            # Insert chart
            worksheet.insert_chart("F2", chart)

    def export_experiment_results(
        self, results: Dict[str, Any], filepath: str, include_metadata: bool = True
    ) -> bool:
        """Export experiment results with structured formatting."""
        try:
            wb = Workbook()

            # Summary sheet
            ws_summary = wb.active
            ws_summary.title = "Summary"

            # Write summary data
            summary_data = []
            for key, value in results.items():
                if isinstance(value, (int, float, str)):
                    summary_data.append([key, value])

            if summary_data:
                ws_summary.append(["Parameter", "Value"])
                for row in summary_data:
                    ws_summary.append(row)

            # Detailed data sheets
            if "data" in results and isinstance(results["data"], list):
                ws_data = wb.create_sheet("Detailed Data")
                df = pd.DataFrame(results["data"])

                # Write headers
                ws_data.append(df.columns.tolist())

                # Write data
                for _, row in df.iterrows():
                    ws_data.append(row.tolist())  # type: ignore

            # Metadata sheet
            if include_metadata:
                ws_meta = wb.create_sheet("Metadata")
                metadata = [
                    ["Export Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
                    ["Export Format", "Excel (.xlsx)"],
                    ["Data Source", "APGI Framework"],
                    ["Version", "1.0"],
                ]

                for row in metadata:
                    ws_meta.append(row)  # type: ignore

            # Save workbook
            wb.save(filepath)
            return True

        except Exception as e:
            print(f"Error exporting experiment results: {e}")
            return False

    def export_multiple_datasets(
        self, datasets: Dict[str, Union[Dict, List, pd.DataFrame]], filepath: str
    ) -> bool:
        """Export multiple datasets to different sheets."""
        try:
            if OPENPYXL_AVAILABLE:
                wb = Workbook()

                # Remove default sheet
                wb.remove(wb.active)

                # Add each dataset as a sheet
                for sheet_name, data in datasets.items():
                    ws = wb.create_sheet(title=sheet_name)

                    # Convert to DataFrame
                    if isinstance(data, dict):
                        df = pd.DataFrame([data])
                    elif isinstance(data, list):
                        df = pd.DataFrame(data)
                    else:
                        df = data

                    # Write headers
                    if not df.empty:
                        ws.append(df.columns.tolist())

                        # Write data
                        for _, row in df.iterrows():
                            ws.append(row.tolist())

                # Save workbook
                wb.save(filepath)
                return True
            else:
                # Fallback to pandas
                with pd.ExcelWriter(filepath) as writer:
                    for sheet_name, data in datasets.items():
                        if isinstance(data, dict):
                            df = pd.DataFrame([data])
                        elif isinstance(data, list):
                            df = pd.DataFrame(data)
                        else:
                            df = data

                        df.to_excel(writer, sheet_name=sheet_name, index=False)
                return True

        except Exception as e:
            print(f"Error exporting multiple datasets: {e}")
            return False

    def create_excel_template(self, template_type: str, filepath: str) -> bool:
        """Create an Excel template for data entry."""
        try:
            if template_type == "experiment":
                return self._create_experiment_template(filepath)
            elif template_type == "parameter":
                return self._create_parameter_template(filepath)
            elif template_type == "results":
                return self._create_results_template(filepath)
            else:
                return False
        except Exception as e:
            print(f"Error creating template: {e}")
            return False

    def _create_experiment_template(self, filepath: str) -> bool:
        """Create experiment configuration template."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Experiment Config"

        # Headers
        headers = [
            "Experiment Name",
            "Description",
            "Number of Participants",
            "Trials per Condition",
            "Session Duration (min)",
            "Parameters",
            "Notes",
        ]
        ws.append(headers)

        # Sample data
        sample_data = [
            ["Interoceptive Gating", "Test interoceptive attention", 20, 50, 60, "", ""]
        ]
        for row in sample_data:
            ws.append(row)

        wb.save(filepath)
        return True

    def _create_parameter_template(self, filepath: str) -> bool:
        """Create parameter configuration template."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Parameters"

        # Headers
        headers = [
            "Parameter Name",
            "Value",
            "Default Value",
            "Min Value",
            "Max Value",
            "Description",
            "Category",
        ]
        ws.append(headers)

        # Sample parameters
        sample_params = [
            [
                "exteroceptive_precision",
                "1.0",
                "1.0",
                "0.1",
                "10.0",
                "Precision of exteroceptive sensory information",
                "APGI Core",
            ],
            [
                "interoceptive_precision",
                "1.0",
                "1.0",
                "0.1",
                "10.0",
                "Precision of interoceptive bodily signals",
                "APGI Core",
            ],
            [
                "somatic_gain",
                "1.0",
                "1.0",
                "0.0",
                "5.0",
                "Gain factor for somatic marker signals",
                "APGI Core",
            ],
        ]

        for param in sample_params:
            ws.append(param)

        wb.save(filepath)
        return True

    def _create_results_template(self, filepath: str) -> bool:
        """Create results recording template."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Results"

        # Headers
        headers = [
            "Experiment ID",
            "Participant ID",
            "Condition",
            "Trial Number",
            "Response",
            "Reaction Time",
            "Accuracy",
            "Confidence",
            "Timestamp",
            "Notes",
        ]
        ws.append(headers)

        wb.save(filepath)
        return True

    def validate_excel_file(self, filepath: str) -> Dict[str, Any]:
        """Validate an Excel file and return validation results."""
        try:
            # Check if file exists
            if not os.path.exists(filepath):
                return {"valid": False, "error": "File does not exist"}

            # Try to read with pandas
            df = pd.read_excel(filepath)

            validation_result = {
                "valid": True,
                "rows": len(df),
                "columns": len(df.columns),
                "column_names": df.columns.tolist(),
                "file_size": os.path.getsize(filepath),
                "warnings": [],
            }

            # Check for common issues
            if len(df) == 0:
                validation_result["warnings"].append("File contains no data rows")  # type: ignore

            if df.isnull().all().all():
                validation_result["warnings"].append("All cells are empty")  # type: ignore

            # Check for duplicate column names
            if len(df.columns) != len(set(df.columns)):
                validation_result["warnings"].append("Duplicate column names found")  # type: ignore

            return validation_result

        except Exception as e:
            return {"valid": False, "error": str(e)}


def get_excel_export_manager() -> ExcelExportManager:
    """Get the Excel export manager instance."""
    if not OPENPYXL_AVAILABLE and not XLSXWRITER_AVAILABLE:
        print(
            "Warning: Neither openpyxl nor xlsxwriter is available. Excel export will be limited to basic pandas functionality."
        )

    return ExcelExportManager()


def export_data_to_excel(
    data: Union[Dict, List, pd.DataFrame], filepath: str, **kwargs
) -> bool:
    """Convenience function to export data to Excel."""
    manager = get_excel_export_manager()
    return manager.export_to_excel(data, filepath, **kwargs)


def create_excel_export_dialog(parent, export_callback) -> None:
    """Create an Excel export dialog."""
    import tkinter as tk
    from tkinter import filedialog, messagebox, ttk

    dialog = tk.Toplevel(parent)
    dialog.title("Export to Excel")
    dialog.geometry("500x400")
    dialog.resizable(False, False)

    # Make dialog modal
    dialog.transient(parent)
    dialog.grab_set()

    # Main frame
    main_frame = ttk.Frame(dialog, padding="20")
    main_frame.pack(fill=tk.BOTH, expand=True)

    # Title
    title_label = ttk.Label(
        main_frame, text="Export to Excel", font=("Arial", 14, "bold")
    )
    title_label.pack(pady=(0, 20))

    # Export options
    options_frame = ttk.LabelFrame(main_frame, text="Export Options", padding="10")
    options_frame.pack(fill=tk.X, pady=(0, 20))

    # Include charts
    include_charts_var = tk.BooleanVar(value=True)
    include_charts_check = ttk.Checkbutton(
        options_frame, text="Include Charts", variable=include_charts_var
    )
    include_charts_check.pack(anchor=tk.W, pady=5)

    # Include styling
    include_styling_var = tk.BooleanVar(value=True)
    include_styling_check = ttk.Checkbutton(
        options_frame, text="Apply Formatting", variable=include_styling_var
    )
    include_styling_check.pack(anchor=tk.W, pady=5)

    # Include metadata
    include_metadata_var = tk.BooleanVar(value=True)
    include_metadata_check = ttk.Checkbutton(
        options_frame, text="Include Metadata", variable=include_metadata_var
    )
    include_metadata_check.pack(anchor=tk.W, pady=5)

    # File selection
    file_frame = ttk.LabelFrame(main_frame, text="Output File", padding="10")
    file_frame.pack(fill=tk.X, pady=(0, 20))

    filepath_var = tk.StringVar()
    filepath_entry = ttk.Entry(file_frame, textvariable=filepath_var, width=50)
    filepath_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 10))

    def browse_file():
        filename = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        )
        if filename:
            filepath_var.set(filename)

    browse_button = ttk.Button(file_frame, text="Browse...", command=browse_file)
    browse_button.pack(side=tk.RIGHT)

    # Buttons
    button_frame = ttk.Frame(main_frame)
    button_frame.pack(pady=(0, 10))

    def export():
        filepath = filepath_var.get()
        if not filepath:
            messagebox.showerror("Error", "Please select an output file.")
            return

        options = {
            "include_charts": include_charts_var.get(),
            "styling": include_styling_var.get(),
            "include_metadata": include_metadata_var.get(),
        }

        if export_callback(filepath, options):
            dialog.destroy()
        else:
            messagebox.showerror("Error", "Failed to export data.")

    export_button = ttk.Button(button_frame, text="Export", command=export)
    export_button.pack(side=tk.LEFT, padx=(0, 10))

    cancel_button = ttk.Button(button_frame, text="Cancel", command=dialog.destroy)
    cancel_button.pack(side=tk.LEFT)

    # Center dialog
    dialog.update_idletasks()
    x = (dialog.winfo_screenwidth() // 2) - (dialog.winfo_width() // 2)
    y = (dialog.winfo_screenheight() // 2) - (dialog.winfo_height() // 2)
    dialog.geometry(f"+{x}+{y}")
